"""v3.8 acceptance checks against the 15-payer exec workbook.

Run after generating the 15-payer output:
    python main.py --mode executive --seed data/seed_payers_15_v38.csv --out out/exec15_v3_8/
    python _acceptance_v3_8.py
"""
import openpyxl
from pathlib import Path

OUT = Path("out/exec15_v3_8")
xlsx = next(p for p in OUT.glob("*.xlsx") if not p.name.startswith("~$"))
wb = openpyxl.load_workbook(xlsx)
ws = wb.active
headers = [c.value for c in ws[1]]
rows = [dict(zip(headers, [c.value for c in r])) for r in ws.iter_rows(min_row=2) if r[0].value]
print(f"TOTAL ROWS: {len(rows)}  FILE: {xlsx.name}")
print("--- ACCEPTANCE CHECKS ---")

# 1. Horizon CEO fix (2-year news window)
horizon_ceo = [r for r in rows if "Horizon" in (r["Payer Name"] or "") and r["Persona"] == "CEO"]
ok1 = bool(horizon_ceo) and "Koenig" not in str(horizon_ceo[0]["Executive Name"] or "")
print(f"1. Horizon CEO != Koenig: {horizon_ceo[0]['Executive Name'] if horizon_ceo else 'MISSING'} -> {'PASS' if ok1 else 'FAIL'}")

# 2. EmblemHealth CEO fix
emblem_ceo = [r for r in rows if "EmblemHealth" in (r["Payer Name"] or "") and r["Persona"] == "CEO"]
ok2 = bool(emblem_ceo) and "Ignagni" not in str(emblem_ceo[0]["Executive Name"] or "")
print(f"2. EmblemHealth CEO != Ignagni: {emblem_ceo[0]['Executive Name'] if emblem_ceo else 'MISSING'} -> {'PASS' if ok2 else 'FAIL'}")

# 3. UPMC CEO fix
upmc_ceo = [r for r in rows if "UPMC" in (r["Payer Name"] or "") and r["Persona"] == "CEO"]
ok3 = bool(upmc_ceo) and "Holder" not in str(upmc_ceo[0]["Executive Name"] or "")
print(f"3. UPMC CEO != Holder: {upmc_ceo[0]['Executive Name'] if upmc_ceo else 'MISSING'} -> {'PASS' if ok3 else 'FAIL'}")

# 4. Excellus CEO fix
excellus_ceo = [r for r in rows if "Excellus" in (r["Payer Name"] or "") and r["Persona"] == "CEO"]
ok4 = bool(excellus_ceo) and "Van Wie" not in str(excellus_ceo[0]["Executive Name"] or "")
print(f"4. Excellus CEO != Van Wie: {excellus_ceo[0]['Executive Name'] if excellus_ceo else 'MISSING'} -> {'PASS' if ok4 else 'FAIL'}")

# 5. Geisinger CEO fix (Direct leadership page fetch)
geisinger_ceo = [r for r in rows if "Geisinger" in (r["Payer Name"] or "") and r["Persona"] == "CEO"]
ok5 = bool(geisinger_ceo) and "Gilliland" in str(geisinger_ceo[0]["Executive Name"] or "")
print(f"5. Geisinger CEO == Gilliland: {geisinger_ceo[0]['Executive Name'] if geisinger_ceo else 'MISSING'} -> {'PASS' if ok5 else 'FAIL'}")

# 6. WellSense CEO fix (Direct leadership page fetch)
wellsense_ceo = [r for r in rows if "WellSense" in (r["Payer Name"] or "") and r["Persona"] == "CEO"]
ok6 = bool(wellsense_ceo) and "Thiltgen" in str(wellsense_ceo[0]["Executive Name"] or "")
print(f"6. WellSense CEO == Thiltgen: {wellsense_ceo[0]['Executive Name'] if wellsense_ceo else 'MISSING'} -> {'PASS' if ok6 else 'FAIL'}")

# 7. UHC CEO Regression
uhc_ceo = [r for r in rows if r["Payer Name"] == "UnitedHealthcare" and r["Persona"] == "CEO"]
ok7 = bool(uhc_ceo) and "Noel" in str(uhc_ceo[0]["Executive Name"] or "")
print(f"7. UHC CEO == Noel: {uhc_ceo[0]['Executive Name'] if uhc_ceo else 'MISSING'} -> {'PASS' if ok7 else 'FAIL'}")

# 8. IBX AmeriHealth Leak Regression
ibx_rows = [r for r in rows if r["Payer Name"] == "Independence Blue Cross"]
amerihealth_leak = [r for r in ibx_rows if any(
    "amerihealth" in str(r.get(col) or "").lower()
    for col in ["Executive Name", "Exact Title"]
)]
ok8 = not amerihealth_leak
print(f"8. IBX no AmeriHealth current-role leak: {len(amerihealth_leak)} leaks -> {'PASS' if ok8 else 'FAIL'}")

# 9. CareOregon Education Filter Regression
co_cio = [r for r in rows if r["Payer Name"] == "CareOregon" and r["Persona"] == "CIO"]
ok9 = True
if co_cio:
    pj_text = (str(co_cio[0]["Past Job 1 Firm"] or "") + " " + str(co_cio[0]["Past Job 2 Firm"] or "")).lower()
    if any(k in pj_text for k in ["university", "college", "school of"]):
        ok9 = False
print(f"9. CareOregon CIO no-edu past_jobs -> {'PASS' if ok9 else 'FAIL'}")

# Summary
passed = sum([ok1, ok2, ok3, ok4, ok5, ok6, ok7, ok8, ok9])
print(f"\n=== RESULT: {passed}/9 checks passed ===")

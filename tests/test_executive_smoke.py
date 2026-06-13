"""Smoke test for the executive pipeline with mocked search + LLM.

Verifies the run_executive() orchestration produces a valid Excel file
without making any network calls.
"""
from __future__ import annotations

import csv
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from payer_intel.schema import EXECUTIVE_EXCEL_COLUMNS


# Mock SearchApi responses keyed by query substring → list of results.
def _fake_google(query: str, num: int = 10, **kwargs):
    q = query.lower()
    if "site:linkedin.com" in q and "chief information officer" in q:
        return [{
            "title": "Jane Doe - Chief Information Officer at Humana",
            "link": "https://www.linkedin.com/in/jane-doe-humana/",
            "snippet": "Chief Information Officer at Humana · 2022 - Present. Previously VP Technology at Anthem.",
            "date": "2026-01-15",
        }]
    if "site:linkedin.com" in q and "chief executive officer" in q:
        return [{
            "title": "Bruce Broussard - CEO at Humana",
            "link": "https://www.linkedin.com/in/bruce-broussard/",
            "snippet": "President & CEO at Humana · 2013 - Present",
            "date": "2026-02-01",
        }]
    if "leadership" in q or "executive team" in q:
        return [{
            "title": "Humana Leadership",
            "link": "https://humana.com/about/leadership",
            "snippet": "Bruce Broussard, President & CEO. Jane Doe, Chief Information Officer.",
            "date": "2026-05-01",
        }]
    return []


def _fake_google_news(query: str, time_range: str = "qdr:y", num: int = 10, **kwargs):
    if "humana" in query.lower():
        return [{
            "title": "Humana appoints Jane Doe as Chief Information Officer",
            "link": "https://businesswire.com/humana-cio-jane-doe",
            "snippet": "Humana today announced the appointment of Jane Doe as CIO, effective immediately.",
            "date": "2026-04-10",
        }]
    return []


def _fake_classify(payer, evidence):
    """Stand-in for _classify_executives_with_llm — returns plausible output."""
    from payer_intel.schema import ExecutiveRole
    if not evidence:
        return {}, "", ""
    return (
        {
            ExecutiveRole.CEO: {
                "name": "Bruce Broussard",
                "title": "President & CEO",
                "current_employer_extracted": "Humana",
                "linkedin_url": "https://www.linkedin.com/in/bruce-broussard/",
                "past_jobs": [],
                "departure_risk": False,
                "departure_note": "",
                "evidence_indices": [
                    i for i, e in enumerate(evidence)
                    if "broussard" in (e.url or "").lower()
                       or "broussard" in (e.snippet or "").lower()
                ],
            },
            ExecutiveRole.CIO: {
                "name": "Jane Doe",
                "title": "Chief Information Officer",
                "current_employer_extracted": "Humana",
                "linkedin_url": "https://www.linkedin.com/in/jane-doe-humana/",
                "past_jobs": [
                    {"firm": "Anthem", "title": "VP Technology", "years": "2022-2024"},
                ],
                "departure_risk": False,
                "departure_note": "",
                "evidence_indices": [
                    i for i, e in enumerate(evidence)
                    if "jane-doe" in (e.url or "").lower()
                       or "jane doe" in (e.snippet or "").lower()
                ],
            },
        },
        "CIO Jane Doe joined recently from Anthem — warm-intro opportunity.",
        "Bruce Broussard remains CEO; Jane Doe confirmed as new CIO via April 2026 press release.",
    )


def test_executive_smoke_run(tmp_path: Path, monkeypatch):
    # Build a tiny 1-payer seed so we don't depend on data/ files.
    seed = tmp_path / "seed.csv"
    with open(seed, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["payer_name", "domain", "payer_type", "search_aliases", "search_excludes"])
        w.writerow(["Humana Inc.", "humana.com", "National", "Humana", ""])

    # Provide a fake API key so SearchApiClient can be instantiated.
    monkeypatch.setenv("SEARCHAPI_API_KEY", "test-key")

    out_dir = tmp_path / "out"

    with patch("payer_intel.crew.SearchApiClient") as mock_client_cls, \
         patch("payer_intel.crew._classify_executives_with_llm", side_effect=_fake_classify), \
         patch("payer_intel.crew._enrich_executive_pages", side_effect=lambda evs, dom: evs):
        client = mock_client_cls.return_value
        client.google.side_effect = _fake_google
        client.google_news.side_effect = _fake_google_news

        from payer_intel.crew import run_executive
        out = run_executive(seed, out_dir)

    assert out.exists()
    wb = load_workbook(out)
    assert "Executive Intelligence" in wb.sheetnames
    ws = wb["Executive Intelligence"]
    assert [c.value for c in ws[1]] == EXECUTIVE_EXCEL_COLUMNS
    # Header + 5 persona rows for 1 payer
    assert ws.max_row == 6
    header = [c.value for c in ws[1]]
    persona_col = header.index("Persona") + 1
    name_col = header.index("Executive Name") + 1

    def _row_for(persona: str) -> int:
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=persona_col).value == persona:
                return r
        raise AssertionError(persona)

    assert ws.cell(row=_row_for("CEO"), column=name_col).value == "Bruce Broussard"
    assert ws.cell(row=_row_for("CIO"), column=name_col).value == "Jane Doe"
    # Unidentified role → placeholder
    assert ws.cell(row=_row_for("CMO"), column=name_col).value == "\u2014"


def test_is_known_deceased_drops_short_surname():
    """v3.2 guard: "Sam Ho" must be dropped despite the 2-char last name
    that would defeat a last-name-only matcher."""
    from payer_intel.crew import _is_known_deceased
    from payer_intel.schema import Evidence

    obit = Evidence(
        source_type="executive_news",
        url="https://example.com/sam-ho-obit",
        snippet="Dr. Sam Ho, longtime Chief Medical Officer at UnitedHealthcare, passed away.",
    )
    living = Evidence(
        source_type="executive_news",
        url="https://example.com/tim-noel",
        snippet="Tim Noel named CEO of UnitedHealthcare.",
    )
    assert _is_known_deceased("Sam Ho", [obit, living]) is True
    assert _is_known_deceased("Tim Noel", [obit, living]) is False
    # Full-name anchor prevents false positives: a different exec with the
    # same surname does not get dropped just because someone else died.
    assert _is_known_deceased("Alex Ho", [obit, living]) is False


def test_is_known_deceased_proximity_avoids_false_positive():
    """v3.5: Tim Noel must NOT be flagged as deceased when an UHC succession
    article mentions both his appointment AND Brian Thompson's death in the
    same document but hundreds of chars apart."""
    from payer_intel.crew import _is_known_deceased
    from payer_intel.schema import Evidence

    # Build a 600+ char body where the name and the deceased signal are
    # ~400 chars apart — well beyond the 150-char proximity window.
    succession_body = (
        "Tim Noel has been appointed Chief Executive Officer of "
        "UnitedHealthcare effective January 2025. Noel previously led the "
        "Medicare & Retirement division and brings two decades of "
        "experience to the role. The company emphasized continuity and a "
        "focus on Medicare Advantage growth. "
        + "Filler content about strategy and Q1 earnings. " * 8
        + "Brian Thompson, the previous CEO, passed away in December 2024."
    )
    ev = Evidence(
        source_type="executive_news",
        url="https://example.com/uhc-noel-ceo",
        full_text=succession_body,
    )
    # Verify the two phrases really are far apart.
    assert succession_body.lower().find("passed away") - succession_body.lower().find(
        "tim noel"
    ) > 200
    assert _is_known_deceased("Tim Noel", [ev]) is False

    # Positive control: a tight obit phrase MUST still flag.
    tight = Evidence(
        source_type="executive_news",
        url="https://example.com/x",
        full_text="Dr. Jane Smith, longtime CMO at Anthem, passed away last week.",
    )
    assert _is_known_deceased("Jane Smith", [tight]) is True


def test_is_known_deceased_rejects_succession_sentence():
    """v3.5.1: Tim Noel must NOT be flagged when the evidence is a
    succession announcement that mentions another exec's death."""
    from payer_intel.crew import _is_known_deceased
    from payer_intel.schema import Evidence

    succession_text = (
        "UnitedHealthcare has appointed Tim Noel as CEO, succeeding Brian "
        "Thompson, who passed away in December 2024 after a tragic incident."
    )
    ev = Evidence(
        source_type="executive_news",
        snippet=succession_text,
        full_text=succession_text,
        url="https://example.com",
        date="2025-01-10",
    )
    assert _is_known_deceased("Tim Noel", [ev]) is False, (
        "Tim Noel must NOT be flagged in a succession announcement "
        "(forward-only window stops at 80 chars after the name)"
    )


def test_is_known_deceased_catches_direct_obituary():
    """v3.5.1: positive control — Sam Ho must still flag when the obit
    text places the deceased verb immediately after his name."""
    from payer_intel.crew import _is_known_deceased
    from payer_intel.schema import Evidence

    obituary_text = (
        "UnitedHealthcare Chief Medical Officer Sam Ho passed away on "
        "January 15, 2026, after a brief illness."
    )
    ev = Evidence(
        source_type="executive_news",
        snippet=obituary_text,
        full_text=obituary_text,
        url="https://example.com",
        date="2026-01-20",
    )
    assert _is_known_deceased("Sam Ho", [ev]) is True


def test_is_known_deceased_rejects_following_tragic_death_of_other():
    """v3.5.1 regression: real WMUR snippet for Tim Noel said he was named
    CEO 'following the tragic death of Brian Thompson'. The deceased signal
    'death of' sat ~30 chars after 'Tim Noel' but referred to Brian. The
    succession-context guard must catch the 'following'/'tragic' deflector
    in between."""
    from payer_intel.crew import _is_known_deceased
    from payer_intel.schema import Evidence

    wmur_snippet = (
        "UnitedHealthcare named Tim Noel as CEO following the tragic death "
        "of Brian Thompson, with increased security measures put in place."
    )
    ev = Evidence(
        source_type="executive_news",
        snippet=wmur_snippet,
        full_text=wmur_snippet,
        url="https://www.wmur.com/article/unitedhealthcare-new-ceo-tim-noel/63532256",
        date="2025-01-10",
    )
    assert _is_known_deceased("Tim Noel", [ev]) is False, (
        "Tim Noel must NOT be flagged when 'death of' refers to his "
        "predecessor Brian Thompson via 'following the tragic death of'"
    )


def test_deduplicate_personas_keeps_higher_priority():
    """v3.5: same exec in two persona slots → keep higher-priority slot,
    clear the other. BCBSK Mike Gerrish (CMO + VP Experience) case."""
    from payer_intel.crew import _deduplicate_personas
    from payer_intel.schema import ExecutiveRole

    classified = {
        ExecutiveRole.CMO: {"name": "Mike Gerrish", "title": "CMO"},
        ExecutiveRole.VP_EXPERIENCE: {
            "name": "Mike Gerrish", "title": "VP Experience",
        },
        ExecutiveRole.CEO: {"name": "Matt All", "title": "President and CEO"},
    }
    out = _deduplicate_personas(classified)
    assert ExecutiveRole.CMO in out  # CMO wins over VPX
    assert ExecutiveRole.VP_EXPERIENCE not in out
    assert ExecutiveRole.CEO in out  # unrelated slot untouched
    # Case-insensitive matching
    classified2 = {
        ExecutiveRole.CMO: {"name": "Mike Gerrish", "title": "CMO"},
        ExecutiveRole.VP_EXPERIENCE: {
            "name": "  mike gerrish  ", "title": "VPX",
        },
    }
    out2 = _deduplicate_personas(classified2)
    assert ExecutiveRole.CMO in out2
    assert ExecutiveRole.VP_EXPERIENCE not in out2


def test_already_departed_clears_slot():
    """v3.3: when departure_note describes a past event (e.g. "departed in
    March 2026"), the slot is blanked but the departure_note survives. A
    future-tense retirement ("retiring end of 2026") keeps the exec in seat."""
    from payer_intel.crew import assemble_executive_record
    from payer_intel.schema import ExecutiveRole

    payer = {"payer_name": "BCBS Louisiana", "payer_type": "Blues Plan", "domain": "bcbsla.com"}
    classified = {
        ExecutiveRole.CEO: {
            "name": "Bryan Camerlinck",
            "title": "President and CEO",
            "linkedin_url": None,
            "past_jobs": [],
            "departure_risk": True,
            "departure_note": "Announced retirement plans for end of 2026.",
            "evidence_indices": [],
        },
        ExecutiveRole.CIO: {
            "name": "Tina Bourgeois",
            "title": "SVP and Chief Information Officer",
            "linkedin_url": None,
            "past_jobs": [],
            "departure_risk": True,
            "departure_note": "Tina Bourgeois has departed in March 2026.",
            "evidence_indices": [],
        },
    }
    rec = assemble_executive_record(payer, classified, [], "", "")
    ceo = rec.executives[ExecutiveRole.CEO]
    cio = rec.executives[ExecutiveRole.CIO]
    # CEO announced future retirement → stays in seat
    assert ceo.name == "Bryan Camerlinck"
    assert ceo.departure_risk is True
    # CIO already departed → blanked, but departure_note survives for context
    assert cio.name is None
    assert cio.title is None
    assert cio.linkedin_url is None
    assert cio.departure_risk is True
    assert cio.departure_note == "Tina Bourgeois has departed in March 2026."
    # BD notes prefixed with [DEPARTED — slot vacant: CIO]
    assert "DEPARTED" in rec.bd_notes
    assert "CIO" in rec.bd_notes


def test_per_exec_bd_notes_not_identical_across_rows():
    """v3.4: BD Notes column must vary per-row. Previously every row in a
    payer block carried the same payer-level summary."""
    from payer_intel.crew import assemble_executive_record
    from payer_intel.export import _exec_record_to_rows
    from payer_intel.schema import ExecutiveRole

    payer = {"payer_name": "Humana", "payer_type": "National", "domain": "humana.com"}
    classified = {
        ExecutiveRole.CEO: {
            "name": "Jim Rechtin", "title": "CEO", "linkedin_url": None,
            "past_jobs": [], "departure_risk": False, "departure_note": None,
            "evidence_indices": [],
            "bd_note": (
                "New CEO joined July 2024 from Envision Healthcare. "
                "Leading post-COVID stabilization. AArete angle: early-tenure "
                "window for strategic advisory engagement."
            ),
        },
        ExecutiveRole.CIO: {
            "name": "Sam Deshpande", "title": "Chief Information Officer",
            "linkedin_url": None, "past_jobs": [], "departure_risk": False,
            "departure_note": None, "evidence_indices": [],
            "bd_note": (
                "CIO since 2021, previously at Wells Fargo. Driving Humana's "
                "cloud-native claims platform. AArete angle: potential fit "
                "for IT cost reduction analytics."
            ),
        },
    }
    rec = assemble_executive_record(payer, classified, [], "", "")
    rows = _exec_record_to_rows(rec)
    bd_values = [r["BD Notes"] for r in rows]

    # 5 rows total, one per persona
    assert len(bd_values) == 5
    # CEO and CIO rows carry their per-exec bd_note (not the payer-level)
    persona_to_bd = {r["Persona"]: r["BD Notes"] for r in rows}
    assert "early-tenure window" in persona_to_bd["CEO"]
    assert "cost reduction analytics" in persona_to_bd["CIO"]
    # Empty CMO / Chief Medical / VP Experience slots get a deterministic
    # template note that mentions the payer + the topic.
    for role_name, topic in (
        ("CMO", "marketing"),
        ("Chief Medical", "clinical"),
        ("VP Experience", "member experience"),
    ):
        assert "Humana" in persona_to_bd[role_name]
        assert topic in persona_to_bd[role_name]
    # Critical assertion: the 5 BD Notes are NOT all identical
    assert len(set(bd_values)) >= 4

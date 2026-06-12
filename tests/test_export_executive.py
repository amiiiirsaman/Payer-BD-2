"""Tests for the executive Excel exporter."""
from pathlib import Path

from openpyxl import load_workbook

from payer_intel.export import write_excel_executive
from payer_intel.schema import (
    EXECUTIVE_EXCEL_COLUMNS,
    ConfidenceScore,
    Evidence,
    ExecutivePayerRecord,
    ExecutiveProfile,
    ExecutiveRole,
)


def _sample_record() -> ExecutivePayerRecord:
    rec = ExecutivePayerRecord(
        payer_name="Humana Inc.",
        payer_type="National",
        domain="humana.com",
        date_verified="2026-06-12",
        confidence=ConfidenceScore.HIGH,
        bd_notes="3/5 roles identified.",
        key_evidence="Strong leadership-page evidence.",
    )
    rec.executives[ExecutiveRole.CEO] = ExecutiveProfile(
        name="Jane Doe",
        title="President & CEO",
        linkedin_url="https://www.linkedin.com/in/jane-doe/",
        past_firms=["Anthem", "UnitedHealth Group"],
        confidence=ConfidenceScore.HIGH,
        evidence=[Evidence(source_type="leadership_page", url="https://humana.com/x")],
    )
    rec.executives[ExecutiveRole.CIO] = ExecutiveProfile(
        name="John Smith",
        title="Chief Information Officer",
        linkedin_url="https://www.linkedin.com/in/john-smith/",
        past_firms=["Accenture"],
        confidence=ConfidenceScore.MEDIUM,
    )
    # CMO, Chief Medical, VP Experience intentionally empty
    rec.executives[ExecutiveRole.CMO] = ExecutiveProfile()
    rec.executives[ExecutiveRole.CHIEF_MEDICAL] = ExecutiveProfile()
    rec.executives[ExecutiveRole.VP_EXPERIENCE] = ExecutiveProfile()
    return rec


def test_export_creates_file_with_three_sheets(tmp_path: Path):
    rec = _sample_record()
    out = write_excel_executive([rec], tmp_path)
    assert out.exists()
    assert out.name.startswith("Aarete_BD_Executive_Intelligence_")
    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Executive Intelligence",
        "Coverage Dashboard",
        "Past Firms Index",
    ]


def test_export_has_16_column_header_in_spec_order(tmp_path: Path):
    out = write_excel_executive([_sample_record()], tmp_path)
    wb = load_workbook(out)
    ws = wb["Executive Intelligence"]
    header = [c.value for c in ws[1]]
    assert header == EXECUTIVE_EXCEL_COLUMNS


def test_export_writes_executive_names_and_hyperlinks(tmp_path: Path):
    out = write_excel_executive([_sample_record()], tmp_path)
    wb = load_workbook(out)
    ws = wb["Executive Intelligence"]
    header = [c.value for c in ws[1]]
    ceo_name_col = header.index("CEO Name") + 1
    ceo_link_col = header.index("CEO LinkedIn") + 1
    assert ws.cell(row=2, column=ceo_name_col).value == "Jane Doe"
    link_cell = ws.cell(row=2, column=ceo_link_col)
    assert link_cell.value == "https://www.linkedin.com/in/jane-doe/"
    assert link_cell.hyperlink is not None
    assert link_cell.hyperlink.target == "https://www.linkedin.com/in/jane-doe/"


def test_export_uses_placeholder_for_missing_executives(tmp_path: Path):
    out = write_excel_executive([_sample_record()], tmp_path)
    wb = load_workbook(out)
    ws = wb["Executive Intelligence"]
    header = [c.value for c in ws[1]]
    cmo_name_col = header.index("CMO/Growth Name") + 1
    assert ws.cell(row=2, column=cmo_name_col).value == "\u2014"  # em-dash placeholder


def test_export_aggregates_past_firms_into_column_m(tmp_path: Path):
    out = write_excel_executive([_sample_record()], tmp_path)
    wb = load_workbook(out)
    ws = wb["Executive Intelligence"]
    header = [c.value for c in ws[1]]
    firms_col = header.index("Past Firms") + 1
    val = ws.cell(row=2, column=firms_col).value or ""
    assert "Anthem" in val
    assert "UnitedHealth Group" in val
    assert "Accenture" in val


def test_coverage_dashboard_counts_identified_vs_missing(tmp_path: Path):
    out = write_excel_executive([_sample_record()], tmp_path)
    wb = load_workbook(out)
    cov = wb["Coverage Dashboard"]
    header = [c.value for c in cov[1]]
    assert header == ["Role", "Identified", "Missing", "High", "Medium", "Low", "Total Payers"]
    # Find the CEO row
    rows = list(cov.iter_rows(min_row=2, values_only=True))
    ceo_row = next(r for r in rows if r[0] == "CEO")
    assert ceo_row[1] == 1  # identified
    assert ceo_row[2] == 0  # missing
    assert ceo_row[3] == 1  # high
    assert ceo_row[6] == 1  # total payers
    cmo_row = next(r for r in rows if r[0] == "CMO")
    assert cmo_row[1] == 0
    assert cmo_row[2] == 1


def test_past_firms_index_lists_each_firm_executive_pair(tmp_path: Path):
    out = write_excel_executive([_sample_record()], tmp_path)
    wb = load_workbook(out)
    sheet = wb["Past Firms Index"]
    header = [c.value for c in sheet[1]]
    assert header == ["Past Firm", "Executive Name", "Current Role", "Current Payer", "LinkedIn"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    # CEO has 2 past firms, CIO has 1 → 3 rows total
    assert len(rows) == 3
    firms = {r[0] for r in rows}
    assert {"Anthem", "UnitedHealth Group", "Accenture"} <= firms


def test_export_handles_empty_record_list(tmp_path: Path):
    out = write_excel_executive([], tmp_path)
    assert out.exists()
    wb = load_workbook(out)
    ws = wb["Executive Intelligence"]
    assert [c.value for c in ws[1]] == EXECUTIVE_EXCEL_COLUMNS
    assert ws.max_row == 1  # only header

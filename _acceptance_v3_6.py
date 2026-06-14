"""v3.6 acceptance checks against the 15-payer exec workbook.

Run after generating the 15-payer output:
    python main.py --mode executive --seed data/seed_payers_15_v36.csv --out out/exec15_v3_6/
    python _acceptance_v3_6.py
"""
import openpyxl
from collections import Counter
from pathlib import Path

OUT = Path("out/exec15_v3_6")
xlsx = next(p for p in OUT.glob("*.xlsx") if not p.name.startswith("~$"))
wb = openpyxl.load_workbook(xlsx)
ws = wb.active
headers = [c.value for c in ws[1]]
rows = [dict(zip(headers, [c.value for c in r])) for r in ws.iter_rows(min_row=2) if r[0].value]
print(f"TOTAL ROWS: {len(rows)}  FILE: {xlsx.name}")
print("--- ACCEPTANCE CHECKS ---")

# ── Regression: Preserved from v3.5.1 ─────────────────────────────────────

# 1. UHC CEO must be Tim Noel (not Brian Thompson — deceased)
uhc_ceo = [r for r in rows if r["Payer Name"] == "UnitedHealthcare" and r["Persona"] == "CEO"]
ok1 = bool(uhc_ceo) and uhc_ceo[0]["Executive Name"] == "Tim Noel"
print(f"1. UHC CEO == Tim Noel: {uhc_ceo[0]['Executive Name'] if uhc_ceo else 'MISSING'} -> {'PASS' if ok1 else 'FAIL'}")

# 2. BCBS Kansas VP Experience must not be Mike Gerrish (dedup'd to CMO)
bcbsk_vpx = [r for r in rows if "Kansas" in (r["Payer Name"] or "") and "Kansas City" not in (r["Payer Name"] or "") and r["Persona"] == "VP Experience"]
ok2 = not bcbsk_vpx or bcbsk_vpx[0]["Executive Name"] != "Mike Gerrish"
print(f"2. BCBSK VPX != Gerrish: {bcbsk_vpx[0]['Executive Name'] if bcbsk_vpx else 'MISSING'} -> {'PASS' if ok2 else 'FAIL'}")

# 3. No non-LinkedIn URLs in LinkedIn column
bad_li = [(r["Payer Name"], r["Persona"], r["LinkedIn"]) for r in rows
          if r["LinkedIn"] and "linkedin.com" not in str(r["LinkedIn"]).lower()
          and r["LinkedIn"] not in ("-", "\u2014")]
ok3 = not bad_li
print(f"3. LinkedIn column clean: {len(bad_li)} bad -> {'PASS' if ok3 else 'FAIL ' + str(bad_li[:3])}")

# 4. CareOregon CIO past_jobs must not include universities
co_cio = [r for r in rows if r["Payer Name"] == "CareOregon" and r["Persona"] == "CIO"]
edu = False
if co_cio:
    pj_text = (str(co_cio[0]["Past Job 1 Firm"] or "") + " " + str(co_cio[0]["Past Job 2 Firm"] or "")).lower()
    edu = any(k in pj_text for k in ["university", "college", "school of"])
    print(f"4. CareOregon CIO no-edu past_jobs -> {'FAIL: edu found' if edu else 'PASS'}")
else:
    print("4. CareOregon CIO MISSING -> N/A")
ok4 = not (co_cio and edu)

# 5. BD Notes must be unique per executive row within each payer
all_unique = True
for payer, group in {p: [r for r in rows if r["Payer Name"] == p] for p in {r["Payer Name"] for r in rows}}.items():
    notes = [r["BD Notes"] for r in group if r["BD Notes"] and r["Executive Name"] not in (None, "-", "\u2014")]
    dupes = [(n, c) for n, c in Counter(notes).items() if c > 1]
    if dupes:
        all_unique = False
        for n, c in dupes:
            print(f"   - DUP in {payer} ({c}x): {n[:90]}...")
ok5 = all_unique
print(f"5. BD Notes unique per row within payer -> {'PASS' if ok5 else 'FAIL'}")

# 6. All 15 payers present, 5 rows each
payers = sorted({r["Payer Name"] for r in rows})
counts = Counter(r["Payer Name"] for r in rows)
ok6 = len(payers) == 15 and all(counts[p] == 5 for p in payers)
print(f"6. 15 payers x 5 rows = 75 total: {len(rows)} rows -> {'PASS' if ok6 else 'FAIL'}")

# ── New v3.6 checks ────────────────────────────────────────────────────────

# 7. Horizon CMO must NOT be Richard Popiel (departed 2012 — date-weighting fix)
horizon_cmo = [r for r in rows if "Horizon" in (r["Payer Name"] or "") and r["Persona"] == "CMO"]
ok7 = not horizon_cmo or "Popiel" not in str(horizon_cmo[0]["Executive Name"] or "")
print(f"7. Horizon CMO != Popiel: {horizon_cmo[0]['Executive Name'] if horizon_cmo else 'MISSING'} -> {'PASS' if ok7 else 'FAIL'}")

# 8. BCBS Michigan CEO must NOT be Daniel Loepp (retired Jan 2025 — date-weighting fix)
mi_ceo = [r for r in rows if "Michigan" in (r["Payer Name"] or "") and r["Persona"] == "CEO"]
ok8 = not mi_ceo or "Loepp" not in str(mi_ceo[0]["Executive Name"] or "")
print(f"8. BCBS Michigan CEO != Loepp: {mi_ceo[0]['Executive Name'] if mi_ceo else 'MISSING'} -> {'PASS' if ok8 else 'FAIL'}")

# 9. Humana CIO must NOT be Kyu Rhee (incorrect role — date-weighting fix)
humana_cio = [r for r in rows if r["Payer Name"] == "Humana Inc." and r["Persona"] == "CIO"]
ok9 = not humana_cio or "Rhee" not in str(humana_cio[0]["Executive Name"] or "")
print(f"9. Humana CIO != Kyu Rhee: {humana_cio[0]['Executive Name'] if humana_cio else 'MISSING'} -> {'PASS' if ok9 else 'FAIL'}")

# 10. Independence Blue Cross must NOT have AmeriHealth in CURRENT role columns only.
#     Past Job columns legitimately reflect biographical history (e.g. Kelly Munson
#     served as President of AmeriHealth Caritas before being named IBX CEO).
ibx_rows = [r for r in rows if r["Payer Name"] == "Independence Blue Cross"]
amerihealth_leak = [r for r in ibx_rows if any(
    "amerihealth" in str(r.get(col) or "").lower()
    for col in ["Executive Name", "Exact Title"]
)]
ok10 = not amerihealth_leak
print(f"10. IBX no AmeriHealth in current role: {len(amerihealth_leak)} leaks -> {'PASS' if ok10 else 'FAIL ' + str(amerihealth_leak[:2])}")

# 11. Point32Health CEO must NOT be Tara Gregorio (MA Senior Care Assoc.) and must NOT
#     be Patrick Gilligan (former Harvard Pilgrim pre-merger CEO). Correct current CEO
#     of the combined entity is Cain Hayes — v3.6 date-weighting should surface him.
p32_ceo = [r for r in rows if r["Payer Name"] == "Point32Health Inc." and r["Persona"] == "CEO"]
bad_p32_names = ("Gregorio", "Gilligan")
ok11 = not p32_ceo or not any(n in str(p32_ceo[0]["Executive Name"] or "") for n in bad_p32_names)
print(f"11. Point32Health CEO != Gregorio/Gilligan: {p32_ceo[0]['Executive Name'] if p32_ceo else 'MISSING'} -> {'PASS' if ok11 else 'FAIL'}")

# 12. Elevance Health CEO must still be Gail Boudreaux (confirm no regression)
elevance_ceo = [r for r in rows if r["Payer Name"] == "Elevance Health" and r["Persona"] == "CEO"]
ok12 = bool(elevance_ceo) and "Boudreaux" in str(elevance_ceo[0]["Executive Name"] or "")
print(f"12. Elevance CEO == Boudreaux: {elevance_ceo[0]['Executive Name'] if elevance_ceo else 'MISSING'} -> {'PASS' if ok12 else 'FAIL'}")

# 13. Kaiser CEO must still be Greg Adams (confirm no regression)
kaiser_ceo = [r for r in rows if "Kaiser" in (r["Payer Name"] or "") and r["Persona"] == "CEO"]
ok13 = bool(kaiser_ceo) and "Adams" in str(kaiser_ceo[0]["Executive Name"] or "")
print(f"13. Kaiser CEO == Adams: {kaiser_ceo[0]['Executive Name'] if kaiser_ceo else 'MISSING'} -> {'PASS' if ok13 else 'FAIL'}")

# 14. Aetna CEO must still be Steve Nelson (confirm no regression)
aetna_ceo = [r for r in rows if r["Payer Name"] == "Aetna" and r["Persona"] == "CEO"]
ok14 = bool(aetna_ceo) and "Nelson" in str(aetna_ceo[0]["Executive Name"] or "")
print(f"14. Aetna CEO == Nelson: {aetna_ceo[0]['Executive Name'] if aetna_ceo else 'MISSING'} -> {'PASS' if ok14 else 'FAIL'}")

# Summary
checks = [ok1, ok2, ok3, ok4, ok5, ok6, ok7, ok8, ok9, ok10, ok11, ok12, ok13, ok14]
passed = sum(checks)
print(f"\n=== RESULT: {passed}/{len(checks)} checks passed ===")

"""v3.7 acceptance checks against the 10-payer exec workbook.

Run after generating the output:
    python main.py --mode executive --seed data/seed_payers_10_v37.csv --out out/exec10_v3_7/
    python _acceptance_v3_7.py
"""
import openpyxl
from pathlib import Path

OUT = Path("out/exec10_v3_7")
xlsx = next(p for p in OUT.glob("*.xlsx") if not p.name.startswith("~$"))
wb = openpyxl.load_workbook(xlsx)
ws = wb.active
headers = [c.value for c in ws[1]]
rows = [dict(zip(headers, [c.value for c in r])) for r in ws.iter_rows(min_row=2) if r[0].value]
print(f"TOTAL ROWS: {len(rows)}  FILE: {xlsx.name}")
print("--- ACCEPTANCE CHECKS ---")

def _pick(payer_pred, persona):
    return [r for r in rows if payer_pred(r["Payer Name"] or "") and r["Persona"] == persona]

# 1. Horizon CEO fix
horizon_ceo = _pick(lambda n: "Horizon" in n, "CEO")
ok1 = bool(horizon_ceo) and "Koenig" not in str(horizon_ceo[0]["Executive Name"] or "")
print(f"1. Horizon CEO != Koenig: {horizon_ceo[0]['Executive Name'] if horizon_ceo else 'MISSING'} -> {'PASS' if ok1 else 'FAIL'}")

# 2. EmblemHealth CEO fix
emblem_ceo = _pick(lambda n: "EmblemHealth" in n, "CEO")
ok2 = bool(emblem_ceo) and "Ignagni" not in str(emblem_ceo[0]["Executive Name"] or "")
print(f"2. EmblemHealth CEO != Ignagni: {emblem_ceo[0]['Executive Name'] if emblem_ceo else 'MISSING'} -> {'PASS' if ok2 else 'FAIL'}")

# 3. UPMC CEO fix
upmc_ceo = _pick(lambda n: "UPMC" in n, "CEO")
ok3 = bool(upmc_ceo) and "Holder" not in str(upmc_ceo[0]["Executive Name"] or "")
print(f"3. UPMC CEO != Holder: {upmc_ceo[0]['Executive Name'] if upmc_ceo else 'MISSING'} -> {'PASS' if ok3 else 'FAIL'}")

# 4. Excellus CEO fix
excellus_ceo = _pick(lambda n: "Excellus" in n, "CEO")
ok4 = bool(excellus_ceo) and "Van Wie" not in str(excellus_ceo[0]["Executive Name"] or "")
print(f"4. Excellus CEO != Van Wie: {excellus_ceo[0]['Executive Name'] if excellus_ceo else 'MISSING'} -> {'PASS' if ok4 else 'FAIL'}")

# 5. Geisinger CEO fix
geisinger_ceo = _pick(lambda n: "Geisinger" in n, "CEO")
ok5 = bool(geisinger_ceo) and "Ryu" not in str(geisinger_ceo[0]["Executive Name"] or "")
print(f"5. Geisinger CEO != Ryu: {geisinger_ceo[0]['Executive Name'] if geisinger_ceo else 'MISSING'} -> {'PASS' if ok5 else 'FAIL'}")

# 6. MVP CEO fix
mvp_ceo = _pick(lambda n: "MVP" in n, "CEO")
ok6 = bool(mvp_ceo) and "Gonick" not in str(mvp_ceo[0]["Executive Name"] or "")
print(f"6. MVP CEO != Gonick: {mvp_ceo[0]['Executive Name'] if mvp_ceo else 'MISSING'} -> {'PASS' if ok6 else 'FAIL'}")

# 7. UHC CEO Regression
uhc_ceo = _pick(lambda n: n == "UnitedHealthcare", "CEO")
ok7 = bool(uhc_ceo) and "Noel" in str(uhc_ceo[0]["Executive Name"] or "")
print(f"7. UHC CEO == Noel: {uhc_ceo[0]['Executive Name'] if uhc_ceo else 'MISSING'} -> {'PASS' if ok7 else 'FAIL'}")

# 8. IBX AmeriHealth leak regression
ibx_rows = [r for r in rows if r["Payer Name"] == "Independence Blue Cross"]
leak = [r for r in ibx_rows if any("amerihealth" in str(r.get(col) or "").lower()
                                   for col in ["Executive Name", "Exact Title"])]
ok8 = not leak
print(f"8. IBX no AmeriHealth current-role leak: {len(leak)} -> {'PASS' if ok8 else 'FAIL'}")

# 9. Output contract: 10 payers x 5 rows = 50
ok9 = len(rows) == 50 and len({r["Payer Name"] for r in rows}) == 10
print(f"9. 10 payers x 5 rows = 50 total: {len(rows)} rows -> {'PASS' if ok9 else 'FAIL'}")

passed = sum([ok1, ok2, ok3, ok4, ok5, ok6, ok7, ok8, ok9])
print(f"\n=== RESULT: {passed}/9 checks passed ===")

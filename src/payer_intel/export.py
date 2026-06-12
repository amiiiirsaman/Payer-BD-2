from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema import (
    EXCEL_COLUMNS,
    EXECUTIVE_EXCEL_COLUMNS,
    EXECUTIVE_ROLE_COLUMNS,
    ExecutivePayerRecord,
    ExecutiveRole,
    PRODUCT_COLUMNS,
    PayerRecord,
)


_VERDICT_STYLES: dict[str, tuple[str, str]] = {
    # verdict -> (fill_hex, font_hex)
    "Yes": ("C6EFCE", "276221"),
    "Likely": ("FFEB9C", "9C5700"),
    "No": ("FFC7CE", "9C0006"),
    "Unknown": ("F2F2F2", "808080"),
}


def _record_to_row(rec: PayerRecord) -> dict[str, str]:
    row = {
        "Payer Name": rec.payer_name,
        "Payer Type": rec.payer_type,
        "Source URLs": "\n".join(dict.fromkeys(rec.source_urls)),
        "Date Identified": rec.date_identified,
        "Confidence Score": rec.confidence.value,
        "BD Notes": rec.bd_notes,
        "Key Evidence": rec.key_evidence,
    }
    for col in PRODUCT_COLUMNS:
        row[col] = rec.verdicts.get(col, "Unknown")
    return row


def write_excel(records: Iterable[PayerRecord], out_dir: Path) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%d")
    path = out_dir / f"Aarete_BD_Salesforce_Payer_Intelligence_{stamp}.xlsx"

    records_list = list(records)
    rows = [_record_to_row(r) for r in records_list]
    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payer Intelligence"
    ws.append(EXCEL_COLUMNS)
    for r in df.itertuples(index=False):
        ws.append(list(r))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, _ in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    widths = {
        "Payer Name": 28, "Payer Type": 14, "Source URLs": 55,
        "Date Identified": 16, "Confidence Score": 18, "BD Notes": 30,
        "Key Evidence": 60,
    }
    for col_idx, name in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 18)

    # Per-cell styling for product verdict columns + wrap_text on data rows
    product_col_indices = {col: EXCEL_COLUMNS.index(col) + 1 for col in PRODUCT_COLUMNS}
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row_num in range(2, len(rows) + 2):
        for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
            ws.cell(row=row_num, column=col_idx).alignment = wrap_align
        for product, col_idx in product_col_indices.items():
            cell = ws.cell(row=row_num, column=col_idx)
            style = _VERDICT_STYLES.get(str(cell.value))
            if style:
                fill_hex, font_hex = style
                cell.fill = PatternFill("solid", fgColor=fill_hex)
                cell.font = Font(color=font_hex, bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # Conditional color scale on Confidence Score (kept for High/Medium/Low)
    if rows:
        conf_idx = EXCEL_COLUMNS.index("Confidence Score") + 1
        col_letter = get_column_letter(conf_idx)
        rng = f"{col_letter}2:{col_letter}{len(rows) + 1}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor="C6EFCE"))
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor="FFEB9C"))
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor="FFC7CE"))
        )

    # Second sheet: Summary Dashboard — product × verdict counts
    summary = wb.create_sheet("Summary Dashboard")
    summary.append(["Salesforce Product", "Yes", "Likely", "No", "Unknown", "Total Payers"])
    for cell in summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    total_payers = len(records_list)
    for product in PRODUCT_COLUMNS:
        counts = {"Yes": 0, "Likely": 0, "No": 0, "Unknown": 0}
        for rec in records_list:
            verdict = rec.verdicts.get(product, "Unknown")
            counts[verdict if verdict in counts else "Unknown"] += 1
        summary.append([product, counts["Yes"], counts["Likely"], counts["No"], counts["Unknown"], total_payers])
    summary.column_dimensions["A"].width = 22
    for letter in ("B", "C", "D", "E", "F"):
        summary.column_dimensions[letter].width = 14
    summary.freeze_panes = "B2"
    # Color the Yes/Likely/Unknown columns lightly for readability
    for row_num in range(2, len(PRODUCT_COLUMNS) + 2):
        summary.cell(row=row_num, column=2).fill = PatternFill("solid", fgColor="E2EFDA")  # Yes
        summary.cell(row=row_num, column=3).fill = PatternFill("solid", fgColor="FFF2CC")  # Likely
        summary.cell(row=row_num, column=5).fill = PatternFill("solid", fgColor="F2F2F2")  # Unknown

    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# Executive Intelligence export (--mode executive)
# ─────────────────────────────────────────────────────────────────────────────
_LINK_FONT = Font(color="0563C1", underline="single")
_PLACEHOLDER = "—"


def _exec_record_to_row(rec: ExecutivePayerRecord) -> dict[str, str]:
    row: dict[str, str] = {
        "Payer Name": rec.payer_name,
        "Payer Type": rec.payer_type,
        "Date Verified": rec.date_verified,
        "Confidence Score": rec.confidence.value,
        "BD Notes": rec.bd_notes,
    }
    for role in ExecutiveRole:
        cols = EXECUTIVE_ROLE_COLUMNS[role]
        profile = rec.executives.get(role)
        if not profile or not profile.name:
            for c in cols:
                row[c] = ""
            continue
        # Identity columns (Name, Title, LinkedIn)
        row[cols[0]] = profile.name or ""
        row[cols[1]] = profile.title or ""
        row[cols[2]] = profile.linkedin_url or ""
        # Past job columns: 2 jobs × (Firm, Title, Years)
        for i in range(2):
            firm_col, title_col, years_col = cols[3 + i * 3], cols[4 + i * 3], cols[5 + i * 3]
            if i < len(profile.past_jobs):
                job = profile.past_jobs[i]
                row[firm_col] = job.firm
                row[title_col] = job.title
                row[years_col] = job.years
            else:
                row[firm_col] = row[title_col] = row[years_col] = ""
    return row


def write_excel_executive(
    records: Iterable[ExecutivePayerRecord], out_dir: Path
) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%d")
    path = out_dir / f"Aarete_BD_Executive_Intelligence_{stamp}.xlsx"

    records_list = list(records)
    rows = [_exec_record_to_row(r) for r in records_list]
    df = pd.DataFrame(rows, columns=EXECUTIVE_EXCEL_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Intelligence"
    ws.append(EXECUTIVE_EXCEL_COLUMNS)
    for r in df.itertuples(index=False):
        ws.append(list(r))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, _ in enumerate(EXECUTIVE_EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "B2"

    widths: dict[str, int] = {
        "Payer Name": 26, "Payer Type": 14,
        "Date Verified": 14, "Confidence Score": 16, "BD Notes": 55,
    }
    for role in ExecutiveRole:
        cols = EXECUTIVE_ROLE_COLUMNS[role]
        widths[cols[0]] = 22   # Name
        widths[cols[1]] = 28   # Title
        widths[cols[2]] = 36   # LinkedIn
        for i in range(2):
            widths[cols[3 + i * 3]] = 22   # Past Firm
            widths[cols[4 + i * 3]] = 26   # Past Title
            widths[cols[5 + i * 3]] = 14   # Past Years
    for col_idx, name in enumerate(EXECUTIVE_EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 18)

    wrap_align = Alignment(wrap_text=True, vertical="top")
    placeholder_font = Font(color="A0A0A0", italic=True)
    linkedin_col_indices = {
        EXECUTIVE_ROLE_COLUMNS[role][2]: EXECUTIVE_EXCEL_COLUMNS.index(EXECUTIVE_ROLE_COLUMNS[role][2]) + 1
        for role in ExecutiveRole
    }
    name_col_indices = {
        EXECUTIVE_ROLE_COLUMNS[role][0]: EXECUTIVE_EXCEL_COLUMNS.index(EXECUTIVE_ROLE_COLUMNS[role][0]) + 1
        for role in ExecutiveRole
    }

    for row_num in range(2, len(rows) + 2):
        for col_idx in range(1, len(EXECUTIVE_EXCEL_COLUMNS) + 1):
            ws.cell(row=row_num, column=col_idx).alignment = wrap_align

        # LinkedIn hyperlinks
        for _link_col, col_idx in linkedin_col_indices.items():
            cell = ws.cell(row=row_num, column=col_idx)
            url = str(cell.value or "").strip()
            if url:
                cell.hyperlink = url
                cell.font = _LINK_FONT
            else:
                cell.value = _PLACEHOLDER
                cell.font = placeholder_font
                cell.alignment = Alignment(
                    wrap_text=True, vertical="top", horizontal="center"
                )

        # Placeholder dash for empty name cells
        for _name_col, col_idx in name_col_indices.items():
            cell = ws.cell(row=row_num, column=col_idx)
            if not str(cell.value or "").strip():
                cell.value = _PLACEHOLDER
                cell.font = placeholder_font
                cell.alignment = Alignment(
                    wrap_text=True, vertical="top", horizontal="center"
                )

    # Conditional formatting on Confidence Score column
    if rows:
        conf_idx = EXECUTIVE_EXCEL_COLUMNS.index("Confidence Score") + 1
        col_letter = get_column_letter(conf_idx)
        rng = f"{col_letter}2:{col_letter}{len(rows) + 1}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor="C6EFCE"))
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor="FFEB9C"))
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor="FFC7CE"))
        )

    # ── Sheet 2: Coverage Dashboard ────────────────────────────────────────
    coverage = wb.create_sheet("Coverage Dashboard")
    coverage.append([
        "Role", "Identified", "Missing", "High", "Medium", "Low", "Total Payers",
    ])
    for cell in coverage[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    total_payers = len(records_list)
    for role in ExecutiveRole:
        identified = 0
        conf_counts = {"High": 0, "Medium": 0, "Low": 0}
        for rec in records_list:
            profile = rec.executives.get(role)
            if profile and profile.name:
                identified += 1
                key = profile.confidence.value
                if key in conf_counts:
                    conf_counts[key] += 1
        coverage.append([
            role.value,
            identified,
            max(total_payers - identified, 0),
            conf_counts["High"],
            conf_counts["Medium"],
            conf_counts["Low"],
            total_payers,
        ])
    coverage.column_dimensions["A"].width = 18
    for letter in ("B", "C", "D", "E", "F", "G"):
        coverage.column_dimensions[letter].width = 14
    coverage.freeze_panes = "B2"
    for row_num in range(2, len(ExecutiveRole) + 2):
        coverage.cell(row=row_num, column=4).fill = PatternFill("solid", fgColor="C6EFCE")  # High
        coverage.cell(row=row_num, column=5).fill = PatternFill("solid", fgColor="FFEB9C")  # Medium
        coverage.cell(row=row_num, column=6).fill = PatternFill("solid", fgColor="FFC7CE")  # Low

    # ── Sheet 3: Past Firms Index (powers Warm Intro Mapper UI) ────────────
    firms_sheet = wb.create_sheet("Past Firms Index")
    firms_sheet.append([
        "Past Firm", "Past Title", "Past Years",
        "Executive Name", "Current Role", "Current Payer", "LinkedIn",
    ])
    for cell in firms_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    firm_rows: list[tuple[str, str, str, str, str, str, str]] = []
    for rec in records_list:
        for role in ExecutiveRole:
            profile = rec.executives.get(role)
            if not (profile and profile.name and profile.past_jobs):
                continue
            for job in profile.past_jobs:
                firm_rows.append((
                    job.firm,
                    job.title,
                    job.years,
                    profile.name,
                    role.value,
                    rec.payer_name,
                    profile.linkedin_url or "",
                ))
    firm_rows.sort(key=lambda t: (t[0].lower(), t[5].lower()))
    for row in firm_rows:
        firms_sheet.append(list(row))
    # Hyperlink the LinkedIn column on this sheet (now column 7)
    for row_num in range(2, len(firm_rows) + 2):
        cell = firms_sheet.cell(row=row_num, column=7)
        url = str(cell.value or "").strip()
        if url:
            cell.hyperlink = url
            cell.font = _LINK_FONT
    firms_sheet.column_dimensions["A"].width = 26
    firms_sheet.column_dimensions["B"].width = 28
    firms_sheet.column_dimensions["C"].width = 14
    firms_sheet.column_dimensions["D"].width = 24
    firms_sheet.column_dimensions["E"].width = 16
    firms_sheet.column_dimensions["F"].width = 28
    firms_sheet.column_dimensions["G"].width = 36
    firms_sheet.freeze_panes = "A2"

    wb.save(path)
    return path

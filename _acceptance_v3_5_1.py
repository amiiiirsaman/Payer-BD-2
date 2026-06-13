"""v3.5.1 acceptance checks against the 10-payer exec workbook."""
import openpyxl
from collections import Counter
from pathlib import Path

OUT = Path("out/exec10_v3_5_1")
xlsx = next(OUT.glob("*.xlsx"))
wb = openpyxl.load_workbook(xlsx)
ws = wb.active
headers = [c.value for c in ws[1]]
rows = [dict(zip(headers, [c.value for c in r])) for r in ws.iter_rows(min_row=2) if r[0].value]
print(f"TOTAL ROWS: {len(rows)}  FILE: {xlsx.name}")
print("--- ACCEPTANCE CHECKS ---")

# 1. Tim Noel UHC CEO
uhc_ceo = [r for r in rows if r["Payer Name"] == "UnitedHealthcare" and r["Persona"] == "CEO"]
ok1 = bool(uhc_ceo) and uhc_ceo[0]["Executive Name"] == "Tim Noel"
print(f"1. UHC CEO == Tim Noel: {uhc_ceo[0]['Executive Name'] if uhc_ceo else 'MISSING'} -> {'PASS' if ok1 else 'FAIL'}")

# 2. BCBSK VPX not Mike Gerrish (dedup'd)
bcbsk_vpx = [r for r in rows if "Kansas" in (r["Payer Name"] or "") and "Kansas City" not in (r["Payer Name"] or "") and r["Persona"] == "VP Experience"]
ok2 = bool(bcbsk_vpx) and bcbsk_vpx[0]["Executive Name"] != "Mike Gerrish"
print(f"2. BCBSK VPX != Gerrish: {bcbsk_vpx[0]['Executive Name'] if bcbsk_vpx else 'MISSING'} -> {'PASS' if ok2 else 'FAIL'}")

# 3. No non-LinkedIn URLs in LinkedIn column
bad = [(r["Payer Name"], r["Persona"], r["LinkedIn"]) for r in rows
       if r["LinkedIn"] and "linkedin.com" not in str(r["LinkedIn"]).lower() and r["LinkedIn"] not in ("-", "\u2014")]
print(f"3. LinkedIn column clean: {len(bad)} bad -> {'PASS' if not bad else 'FAIL ' + str(bad)}")

# 4. CareOregon CIO past_jobs (no universities)
co_cio = [r for r in rows if r["Payer Name"] == "CareOregon" and r["Persona"] == "CIO"]
if co_cio:
    pj_text = (str(co_cio[0]["Past Job 1 Firm"] or "") + " " + str(co_cio[0]["Past Job 2 Firm"] or "")).lower()
    edu = any(k in pj_text for k in ["university", "college", "school of"])
    print(f"4. CareOregon CIO past_jobs no-edu: PJ1={co_cio[0]['Past Job 1 Firm']!r} PJ2={co_cio[0]['Past Job 2 Firm']!r} -> {'FAIL: edu' if edu else 'PASS'}")
else:
    print("4. CareOregon CIO MISSING -> N/A")

# 5. BCBSLA CMO != Brian Keller
la_cmo = [r for r in rows if "Louisiana" in (r["Payer Name"] or "") and r["Persona"] == "CMO"]
ok5 = bool(la_cmo) and la_cmo[0]["Executive Name"] != "Brian Keller"
print(f"5. BCBSLA CMO != Brian Keller: {la_cmo[0]['Executive Name'] if la_cmo else 'MISSING'} -> {'PASS' if ok5 else 'FAIL'}")

# 6. Per-exec BD notes unique within payer (when multiple non-empty rows exist)
all_unique = True
for payer, group in {p: [r for r in rows if r["Payer Name"] == p] for p in {r["Payer Name"] for r in rows}}.items():
    notes = [r["BD Notes"] for r in group if r["BD Notes"] and r["Executive Name"] not in (None, "-", "\u2014")]
    dupes = [(n, c) for n, c in Counter(notes).items() if c > 1]
    if dupes:
        all_unique = False
        for n, c in dupes:
            print(f"   - DUP in {payer} ({c}x): {n[:90]}...")
print(f"6. BD Notes unique per row within payer -> {'PASS' if all_unique else 'FAIL'}")

# 7. All 10 payers present, 5 rows each
payers = sorted({r["Payer Name"] for r in rows})
counts = Counter(r["Payer Name"] for r in rows)
print(f"7. Payers ({len(payers)}): " + ", ".join(f"{p}={counts[p]}" for p in payers))
